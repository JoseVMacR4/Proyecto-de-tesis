from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.utils.formats import date_format
from django.utils import timezone
from django.db.models import Q
from django.core.exceptions import ValidationError
from decimal import Decimal
from apps.reconciliation.models import BankStatementTransaction
from apps.users.permissions import can_upload_statements, can_reconcile
from apps.bank_accounts.models import BankAccount, Office, Operation
from apps.users.models import Notification, UserActivity
import json
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Create your views here.
@login_required
def reconciliation_view(request):
	"""
	Vista principal para la conciliación bancaria
	"""
	page_number = request.GET.get('page', 1)
	transactions = BankStatementTransaction.objects.select_related('bank_account', 'bank_statement').order_by('-bank_statement__statement_date', 'current_reference')
	paginator = Paginator(transactions, 50)
	page_obj = paginator.get_page(page_number)

	page_range = []
	total_pages = paginator.num_pages
	current_page = page_obj.number
	window = 5
	if total_pages <= window:
		page_range = range(1, total_pages + 1)
	else:
		left = max(current_page - 2, 1)
		right = min(current_page + 2, total_pages)
		if left == 1:
			right = 5
		elif right == total_pages:
			left = total_pages - 4
		page_range = range(left, right + 1)

	# Calcular estadísticas
	total_transactions = BankStatementTransaction.objects.count()
	reconciled_count = BankStatementTransaction.objects.filter(status=BankStatementTransaction.StatusChoices.RECONCILED).count()
	pending_count = total_transactions - reconciled_count
	
	# Calcular porcentaje de conciliadas
	reconciled_percentage = 0
	if total_transactions > 0:
		reconciled_percentage = round((reconciled_count / total_transactions) * 100)

	context = {
		'title': 'Conciliación Bancaria',
		'page': 'reconciliation',
		'current_page': 'reconciliation',
		'transactions_page': page_obj,
		'page_range': page_range,
		'total_transactions': paginator.count,
		'total_count': total_transactions,
		'reconciled_count': reconciled_count,
		'pending_count': pending_count,
		'reconciled_percentage': reconciled_percentage,
	}
	return render(request, 'reconciliation/reconciliation.html', context)

@require_http_methods(["GET"])
@login_required
def get_reconciliation_data(request):
	try:
		date_range = request.GET.get('date_range', '')
		date_from = request.GET.get('date_from', '')
		date_to = request.GET.get('date_to', '')
		bank = request.GET.get('bank', '')
		status = request.GET.get('status', '')
		reference = request.GET.get('reference', '')
		name = request.GET.get('name', '')
		office_code = request.GET.get('office_code', '')
		entry_type = request.GET.get('entry_type', '')
		operation_type = request.GET.get('operation_type', '')
		amount_min = request.GET.get('amount_min', '')
		amount_max = request.GET.get('amount_max', '')
		currency = request.GET.get('currency', '')
		page = request.GET.get('page', 1)
		sort_column = request.GET.get('sort', '')
		sort_order = request.GET.get('order', 'asc')

		transactions = BankStatementTransaction.objects.select_related('bank_account', 'bank_statement')

		# Ordenamiento dinámico
		sort_mapping = {
			'date': 'bank_statement__statement_date',
			'operations': 'operation_count',
			'amount': 'amount',
		}

		if sort_column and sort_column in sort_mapping:
			order_prefix = '-' if sort_order == 'desc' else ''
			transactions = transactions.order_by(f"{order_prefix}{sort_mapping[sort_column]}")
		else:
			transactions = transactions.order_by('-bank_statement__statement_date')

		# Filtro por rango de fechas
		if date_range:
			parts = [part.strip() for part in date_range.split('-')]
			if len(parts) == 2:
				try:
					start_date = timezone.datetime.strptime(parts[0], '%d/%m/%Y').date()
					end_date = timezone.datetime.strptime(parts[1], '%d/%m/%Y').date()
					transactions = transactions.filter(bank_statement__statement_date__range=(start_date, end_date))
				except ValueError:
					pass
		elif date_from or date_to:
			if date_from:
				try:
					start_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
					transactions = transactions.filter(bank_statement__statement_date__gte=start_date)
				except ValueError:
					pass
			if date_to:
				try:
					end_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
					transactions = transactions.filter(bank_statement__statement_date__lte=end_date)
				except ValueError:
					pass

		# Filtro por banco (múltiple) - buscar por código
		if bank:
			bank_list = [b.strip() for b in bank.split(',') if b.strip()]
			if bank_list:
				transactions = transactions.filter(bank_account__code__in=bank_list)

		# Filtro por estado (múltiple)
		if status:
			status_list = [s.strip() for s in status.split(',') if s.strip()]
			if status_list:
				transactions = transactions.filter(status__in=status_list)

		# Filtro por referencia
		if reference:
			transactions = transactions.filter(
				Q(current_reference__icontains=reference) | Q(original_reference__icontains=reference)
			)

		# Filtro por nombre/descripción
		if name:
			transactions = transactions.filter(name__icontains=name)

		# Filtro por código de oficina (múltiple)
		if office_code:
			office_list = [o.strip() for o in office_code.split(',') if o.strip()]
			if office_list:
				transactions = transactions.filter(office_code__in=office_list)

		# Filtro por tipo de entrada (múltiple)
		if entry_type:
			entry_list = [e.strip() for e in entry_type.split(',') if e.strip()]
			if entry_list:
				transactions = transactions.filter(entry_type__in=entry_list)

		# Filtro por tipo de operación (múltiple)
		if operation_type:
			operation_list = [o.strip() for o in operation_type.split(',') if o.strip()]
			if operation_list:
				transactions = transactions.filter(operation_type__in=operation_list)

		# Filtro por rango de monto
		if amount_min:
			try:
				min_amount = Decimal(amount_min)
				transactions = transactions.filter(amount__gte=min_amount)
			except:
				pass
		if amount_max:
			try:
				max_amount = Decimal(amount_max)
				transactions = transactions.filter(amount__lte=max_amount)
			except:
				pass

# Filtro por moneda (múltiple)
		if currency:
			currency_list = [c.strip() for c in currency.split(',') if c.strip()]
			if currency_list:
				transactions = transactions.filter(currency__in=currency_list)

		paginator = Paginator(transactions, 50)
		page_obj = paginator.get_page(page)

		office_map = {o.code: o.name for o in Office.objects.all()}
		operation_map = {o.code: o.name for o in Operation.objects.all()}

		transactions_data = []
		for tx in page_obj.object_list:
			try:
				office_value = tx.office_code or ''
				office_display = office_map.get(office_value, office_value) if office_value else ''
				operation_value = tx.operation_type or ''
				operation_display = operation_map.get(operation_value, '') if operation_value else ''
				transactions_data.append({
					'id': str(tx.id),
					'date': date_format(tx.date, 'd/m/Y'),
					'reference': tx.current_reference,
					'original_reference': tx.original_reference,
					'bank': tx.bank_account.name if tx.bank_account else 'N/A',
					'bank_code': tx.bank_account.code if tx.bank_account else 'N/A',
					'office': office_display,
					'office_code': office_value,
					'operations': tx.operation_count,
'description': tx.name,
					'amount': float(tx.amount),
					'entry_type': tx.entry_type,
					'bank_fee': float(tx.bank_fee),
					'operation_type': operation_display,
					'operation_type_code': operation_value,
					'currency': tx.currency,
					'statement_date': date_format(tx.bank_statement.statement_date, 'd/m/Y') if tx.bank_statement else '',
					'status': tx.status,
					'status_display': tx.get_status_display(),
      				'created_at': date_format(tx.created_at, 'd/m/Y H:i') if tx.created_at else '',
					'updated_at': date_format(tx.updated_at, 'd/m/Y H:i') if tx.updated_at else '',
				})
			except Exception as e:
				print(f"Error procesando transacción {tx.id}: {str(e)}")
				continue

		return JsonResponse({
			'status': 'success',
			'page': page_obj.number,
			'total_pages': paginator.num_pages,
			'total_transactions': paginator.count,
			'transactions': transactions_data,
   			'start_index': page_obj.start_index() if paginator.count > 0 else 0,
            'end_index': page_obj.end_index() if paginator.count > 0 else 0,
		})
	except Exception as e:
		import traceback
		print(f"Error en get_reconciliation_data: {str(e)}")
		traceback.print_exc()
		Notification.objects.create(user=request.user, type='error', content=f"Error al obtener datos de conciliación: {str(e)}")
		return JsonResponse({
			'status': 'error',
			'message': str(e)
		}, status=500)

@require_http_methods(["POST"])
@login_required
def reconcile_transaction(request):
    if not can_reconcile(request.user):
        return JsonResponse({'status': 'error', 'message': 'No tiene permiso para conciliar transacciones.'}, status=403)
    
    try:
        data = json.loads(request.body)
        tx_id = data.get('transaction_id')
        
        # Validación preventiva para evitar el error 400 de formato
        if not tx_id or tx_id == "undefined" or tx_id == "null":
            return JsonResponse({
                'status': 'error', 
                'message': 'ID de transacción no válido o vacío'
            }, status=400)
        
        try:
            transaction = BankStatementTransaction.objects.get(id=tx_id)
            transaction.status = BankStatementTransaction.StatusChoices.RECONCILED
            transaction.save()

            Notification.objects.create(
                user=request.user,
                type='info',
                content=f"Transacción {transaction.current_reference} conciliada correctamente."
            )

            UserActivity.objects.create(
                user=request.user,
                action=UserActivity.ActionType.CONCILIATION,
                description=f"Transacción {transaction.current_reference} conciliada",
                metadata={'transaction_id': str(tx_id), 'amount': str(transaction.amount), 'currency': transaction.currency}
            )

            updated_at = date_format(transaction.updated_at, 'd/m/Y H:i')

            return JsonResponse({
                'status': 'success',
                'message': f"Transacción {transaction.current_reference} conciliada correctamente", # Opcional: mejorar el mensaje de respuesta
                'updated_at': updated_at,
            })
        except BankStatementTransaction.DoesNotExist:
            Notification.objects.create(user=request.user, type='error', content="La transacción ya no existe")
            return JsonResponse({'status': 'error', 'message': 'La transacción ya no existe'}, status=404)
        except ValidationError:
            Notification.objects.create(user=request.user, type='error', content="Formato UUID inválido para transacción")
            return JsonResponse({'status': 'error', 'message': 'Formato UUID inválido'}, status=400)

    except Exception as e:
        Notification.objects.create(user=request.user, type='error', content=f"Error al conciliar transacción: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@require_http_methods(["POST"])
@login_required
def reconcile_transactions_bulk(request):
    if not can_reconcile(request.user):
        return JsonResponse({'status': 'error', 'message': 'No tiene permiso para conciliar transacciones.'}, status=403)
    
    try:
        data = json.loads(request.body)
        transaction_ids = data.get('transaction_ids', [])
        
        if not transaction_ids:
            return JsonResponse({'status': 'error', 'message': 'No se proporcionaron transacciones.'}, status=400)
        
        if len(transaction_ids) > 50:
            return JsonResponse({'status': 'error', 'message': 'El límite máximo es de 50 transacciones por operación.'}, status=400)
        
        pending_transactions = BankStatementTransaction.objects.filter(
            id__in=transaction_ids,
            status=BankStatementTransaction.StatusChoices.PENDING
        )
        
        count = pending_transactions.count()
        
        if count == 0:
            return JsonResponse({'status': 'error', 'message': 'No se encontraron transacciones pendientes para conciliar.'}, status=400)
        
        pending_transactions.update(status=BankStatementTransaction.StatusChoices.RECONCILED)
        
        notified_transactions = BankStatementTransaction.objects.filter(
            id__in=transaction_ids,
            status=BankStatementTransaction.StatusChoices.RECONCILED
        )
        
        notifications_to_create = [
            Notification(
                user=request.user,
                type='info',
                content=f"Transacción {tx.current_reference} conciliada correctamente."
            )
            for tx in notified_transactions
        ]
        Notification.objects.bulk_create(notifications_to_create)

        UserActivity.objects.create(
            user=request.user,
            action=UserActivity.ActionType.CONCILIATION,
            description=f"{count} transacción(es) conciliada(s) en lote",
            metadata={'transaction_count': count, 'transaction_ids': [str(id) for id in transaction_ids]}
        )

        return JsonResponse({
            'status': 'success',
            'message': f'{count} transacción(es) conciliada(s) correctamente.',
            'count': count,
        })
    except json.JSONDecodeError:
        Notification.objects.create(user=request.user, type='error', content="Error al conciliar transacciones: formato JSON inválido")
        return JsonResponse({'status': 'error', 'message': 'Formato JSON inválido.'}, status=400)
    except Exception as e:
        Notification.objects.create(user=request.user, type='error', content=f"Error al conciliar transacciones en masa: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@require_http_methods(["POST"])
@login_required
def export_reconciliation(request):
	"""
	API para exportar datos de reconciliación
	"""
	try:
		data = json.loads(request.body)
		format_type = data.get('format', 'xls')
		
		# Aquí se generaría el archivo
		response = {
			'status': 'success',
			'message': 'Archivo generado correctamente',
			'format': format_type
		}
		
		return JsonResponse(response)
	except Exception as e:
		Notification.objects.create(user=request.user, type='error', content=f"Error al exportar conciliación: {str(e)}")
		return JsonResponse({
			'status': 'error',
			'message': str(e)
		}, status=400)

@require_http_methods(["GET"])
@login_required
def get_filter_options(request):
	"""
	API para obtener las opciones disponibles para los filtros
	"""
	try:
		# Obtener bancos
		banks = BankAccount.objects.values('code', 'name').order_by('code')
		bank_options = [{'value': '', 'label': 'Todos los Bancos'}]
		bank_options.extend([
			{'value': bank['code'], 'label': f"{bank['code']} - {bank['name']}"}
			for bank in banks
		])

		# Obtener tipos de entrada únicos
		entry_types = BankStatementTransaction.objects.values_list('entry_type', flat=True).distinct().exclude(entry_type__isnull=True).exclude(entry_type='').order_by('entry_type')
		entry_type_options = [{'value': '', 'label': 'Todos los Tipos'}]
		entry_type_options.extend([
			{'value': et, 'label': 'Crédito (Cr)' if et == 'Cr' else 'Débito (Db)'}
			for et in entry_types
		])

		# Obtener monedas únicas
		currencies = BankStatementTransaction.objects.values_list('currency', flat=True).distinct().exclude(currency__isnull=True).exclude(currency='').order_by('currency')
		currency_options = [{'value': '', 'label': 'Todas las Monedas'}]
		currency_options.extend([
			{'value': curr, 'label': curr}
			for curr in currencies
		])

		# Obtener oficinas registradas y códigos de transacciones
		registered_offices = {o.code: o.name for o in Office.objects.all()}
		transaction_office_codes = BankStatementTransaction.objects.values_list('office_code', flat=True).distinct().exclude(office_code__isnull=True).exclude(office_code='').order_by('office_code')
		
		office_options = [{'value': '', 'label': 'Todas las Oficinas'}]
		for code in transaction_office_codes:
			name = registered_offices.get(code, code)
			office_options.append({
				'value': code,
				'label': f"{code} - {name}" if code != name else code
			})

		# Obtener operaciones registradas y códigos de transacciones
		registered_ops = {o.code: o.name for o in Operation.objects.all()}
		transaction_op_codes = BankStatementTransaction.objects.values_list('operation_type', flat=True).distinct().exclude(operation_type__isnull=True).exclude(operation_type='').order_by('operation_type')
		
		operation_type_options = [{'value': '', 'label': 'Todos los Tipos'}]
		for code in transaction_op_codes:
			name = registered_ops.get(code, code)
			operation_type_options.append({
				'value': code,
				'label': f"{code} - {name}" if code != name else code
			})

		return JsonResponse({
			'status': 'success',
			'filters': {
				'banks': bank_options,
				'entry_types': entry_type_options,
				'currencies': currency_options,
				'offices': office_options,
				'operation_types': operation_type_options,
			}
		})
	except Exception as e:
		return JsonResponse({
			'status': 'error',
			'message': str(e)
		}, status=500)

@require_http_methods(["GET"])
@login_required
def get_reconciliation_stats(request):
    try:
        transactions = BankStatementTransaction.objects.select_related('bank_account', 'bank_statement')

        date_range = request.GET.get('date_range', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        bank = request.GET.get('bank', '')
        status = request.GET.get('status', '')
        reference = request.GET.get('reference', '')
        name = request.GET.get('name', '')
        office_code = request.GET.get('office_code', '')
        entry_type = request.GET.get('entry_type', '')
        operation_type = request.GET.get('operation_type', '')
        amount_min = request.GET.get('amount_min', '')
        amount_max = request.GET.get('amount_max', '')
        currency = request.GET.get('currency', '')

        if date_range:
            parts = [part.strip() for part in date_range.split('-')]
            if len(parts) == 2:
                try:
                    start_date = timezone.datetime.strptime(parts[0], '%d/%m/%Y').date()
                    end_date = timezone.datetime.strptime(parts[1], '%d/%m/%Y').date()
                    transactions = transactions.filter(bank_statement__statement_date__range=(start_date, end_date))
                except ValueError:
                    pass
        elif date_from or date_to:
            if date_from:
                try:
                    start_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
                    transactions = transactions.filter(bank_statement__statement_date__gte=start_date)
                except ValueError:
                    pass
            if date_to:
                try:
                    end_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
                    transactions = transactions.filter(bank_statement__statement_date__lte=end_date)
                except ValueError:
                    pass

        if bank:
            bank_list = [b.strip() for b in bank.split(',') if b.strip()]
            if bank_list:
                transactions = transactions.filter(bank_account__code__in=bank_list)

        if status:
            status_list = [s.strip() for s in status.split(',') if s.strip()]
            if status_list:
                transactions = transactions.filter(status__in=status_list)

        if reference:
            transactions = transactions.filter(
                Q(current_reference__icontains=reference) | Q(original_reference__icontains=reference)
            )

        if name:
            transactions = transactions.filter(name__icontains=name)

        if office_code:
            office_list = [o.strip() for o in office_code.split(',') if o.strip()]
            if office_list:
                transactions = transactions.filter(office_code__in=office_list)

        if entry_type:
            entry_list = [e.strip() for e in entry_type.split(',') if e.strip()]
            if entry_list:
                transactions = transactions.filter(entry_type__in=entry_list)

        if operation_type:
            operation_list = [o.strip() for o in operation_type.split(',') if o.strip()]
            if operation_list:
                transactions = transactions.filter(operation_type__in=operation_list)

        if amount_min:
            try:
                min_amount = Decimal(amount_min)
                transactions = transactions.filter(amount__gte=min_amount)
            except:
                pass
        if amount_max:
            try:
                max_amount = Decimal(amount_max)
                transactions = transactions.filter(amount__lte=max_amount)
            except:
                pass

        if currency:
            currency_list = [c.strip() for c in currency.split(',') if c.strip()]
            if currency_list:
                transactions = transactions.filter(currency__in=currency_list)

        total_transactions = transactions.count()
        reconciled_count = transactions.filter(status=BankStatementTransaction.StatusChoices.RECONCILED).count()
        pending_count = transactions.filter(status=BankStatementTransaction.StatusChoices.PENDING).count()
        
        reconciled_percentage = 0
        if total_transactions > 0:
            reconciled_percentage = round((reconciled_count / total_transactions) * 100)

        return JsonResponse({
            'status': 'success',
            'total_count': total_transactions,
            'reconciled_count': reconciled_count,
            'pending_count': pending_count,
            'reconciled_percentage': reconciled_percentage,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["GET"])
@login_required
def get_export_data(request):
    """
    API para obtener los datos de exportación filtrados
    """
    try:
        date_range = request.GET.get('date_range', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        bank = request.GET.get('bank', '')
        status = request.GET.get('status', '')
        reference = request.GET.get('reference', '')
        name = request.GET.get('name', '')
        office_code = request.GET.get('office_code', '')
        entry_type = request.GET.get('entry_type', '')
        operation_type = request.GET.get('operation_type', '')
        amount_min = request.GET.get('amount_min', '')
        amount_max = request.GET.get('amount_max', '')
        currency = request.GET.get('currency', '')
        selected_ids = request.GET.get('ids', '')

        transactions = BankStatementTransaction.objects.select_related('bank_account', 'bank_statement').order_by('-bank_statement__statement_date', 'current_reference')

        # Aplicar filtros
        if date_range:
            parts = [part.strip() for part in date_range.split('-')]
            if len(parts) == 2:
                try:
                    start_date = timezone.datetime.strptime(parts[0], '%d/%m/%Y').date()
                    end_date = timezone.datetime.strptime(parts[1], '%d/%m/%Y').date()
                    transactions = transactions.filter(bank_statement__statement_date__range=(start_date, end_date))
                except ValueError:
                    pass
        elif date_from or date_to:
            if date_from:
                try:
                    start_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
                    transactions = transactions.filter(bank_statement__statement_date__gte=start_date)
                except ValueError:
                    pass
            if date_to:
                try:
                    end_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
                    transactions = transactions.filter(bank_statement__statement_date__lte=end_date)
                except ValueError:
                    pass

        if bank:
            bank_list = [b.strip() for b in bank.split(',') if b.strip()]
            if bank_list:
                transactions = transactions.filter(bank_account__code__in=bank_list)

        if status:
            status_list = [s.strip() for s in status.split(',') if s.strip()]
            if status_list:
                transactions = transactions.filter(status__in=status_list)

        if reference:
            transactions = transactions.filter(
                Q(current_reference__icontains=reference) | Q(original_reference__icontains=reference)
            )

        if name:
            transactions = transactions.filter(name__icontains=name)

        if office_code:
            office_list = [o.strip() for o in office_code.split(',') if o.strip()]
            if office_list:
                transactions = transactions.filter(office_code__in=office_list)

        if entry_type:
            entry_list = [e.strip() for e in entry_type.split(',') if e.strip()]
            if entry_list:
                transactions = transactions.filter(entry_type__in=entry_list)

        if operation_type:
            operation_list = [o.strip() for o in operation_type.split(',') if o.strip()]
            if operation_list:
                transactions = transactions.filter(operation_type__in=operation_list)

        if amount_min:
            try:
                min_amount = Decimal(amount_min)
                transactions = transactions.filter(amount__gte=min_amount)
            except:
                pass
        if amount_max:
            try:
                max_amount = Decimal(amount_max)
                transactions = transactions.filter(amount__lte=max_amount)
            except:
                pass

        if currency:
            currency_list = [c.strip() for c in currency.split(',') if c.strip()]
            if currency_list:
                transactions = transactions.filter(currency__in=currency_list)

        # Si hay IDs seleccionados, filtrar por esos
        if selected_ids:
            id_list = [id.strip() for id in selected_ids.split(',') if id.strip()]
            transactions = transactions.filter(id__in=id_list)

        office_map = {o.code: o.name for o in Office.objects.all()}
        operation_map = {o.code: o.name for o in Operation.objects.all()}

        transactions_data = []
        total_amount = 0
        for tx in transactions:
            try:
                office_value = tx.office_code or ''
                office_display = office_map.get(office_value, office_value) if office_value else ''
                operation_value = tx.operation_type or ''
                operation_display = operation_map.get(operation_value, '') if operation_value else ''
                total_amount += float(tx.amount)
                transactions_data.append({
                    'id': str(tx.id),
                    'date': date_format(tx.date, 'd/m/Y'),
                    'reference': tx.current_reference,
                    'original_reference': tx.original_reference or '',
                    'bank': tx.bank_account.name if tx.bank_account else 'N/A',
                    'bank_code': tx.bank_account.code if tx.bank_account else 'N/A',
                    'office': office_display,
                    'office_code': office_value,
                    'operations': tx.operation_count,
                    'description': tx.name,
                    'amount': float(tx.amount),
                    'entry_type': tx.entry_type,
                    'bank_fee': float(tx.bank_fee),
                    'operation_type': operation_display,
                    'operation_type_code': operation_value,
                    'currency': tx.currency,
                    'statement_date': date_format(tx.bank_statement.statement_date, 'd/m/Y') if tx.bank_statement else '',
                    'status': tx.status,
                    'status_display': tx.get_status_display(),
                    'created_at': date_format(tx.created_at, 'd/m/Y H:i') if tx.created_at else '',
                    'updated_at': date_format(tx.updated_at, 'd/m/Y H:i') if tx.updated_at else '',
                })
            except Exception as e:
                print(f"Error procesando transacción {tx.id}: {str(e)}")
                continue

        return JsonResponse({
            'status': 'success',
            'count': len(transactions_data),
            'total_amount': total_amount,
            'transactions': transactions_data,
            'filters_applied': {
                'date_range': date_range,
                'date_from': date_from,
                'date_to': date_to,
                'bank': bank,
                'status': status,
                'reference': reference,
                'name': name,
                'office_code': office_code,
                'entry_type': entry_type,
                'operation_type': operation_type,
                'amount_min': amount_min,
                'amount_max': amount_max,
                'currency': currency,
            }
        })
    except Exception as e:
        import traceback
        print(f"Error en get_export_data: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["POST"])
@login_required
def export_download(request):
    """
    API para generar y descargar archivos exportables (XLS, PDF)
    """
    try:
        data = json.loads(request.body)
        format_type = data.get('format', 'xls')
        export_data = data.get('data', {})
        
        transactions = export_data.get('transactions', [])
        filters = export_data.get('filters_applied', {})
        
        if not transactions:
            return JsonResponse({'status': 'error', 'message': 'No hay datos para exportar'}, status=400)
        
        if format_type == 'xls':
            return export_to_xls(transactions, filters)
        elif format_type == 'pdf':
            return export_to_pdf(transactions, filters)
        else:
            return JsonResponse({'status': 'error', 'message': 'Formato no soportado'}, status=400)
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def export_to_xls(transactions, filters):
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'status': 'error', 'message': 'Librería openpyxl no disponible'}, status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        'Fecha', 'Ref. Corriente', 'Ref. Original', 'Banco', 'Oficina',
        'Tipo de Operación', 'Cant. Oper.', 'Descripción', 'Monto', 'Moneda',
        'Comisión', 'Tipo', 'Estado', 'Fecha Creación'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    total_amount = 0
    for row_idx, tx in enumerate(transactions, 2):
        total_amount += tx.get('amount', 0)
        
        ws.cell(row=row_idx, column=1, value=tx.get('date', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=tx.get('reference', '')).border = thin_border
        ws.cell(row=row_idx, column=3, value=tx.get('original_reference', '')).border = thin_border
        ws.cell(row=row_idx, column=4, value=tx.get('bank', '')).border = thin_border
        ws.cell(row=row_idx, column=5, value=tx.get('office', '')).border = thin_border
        ws.cell(row=row_idx, column=6, value=tx.get('operation_type', '')).border = thin_border
        ws.cell(row=row_idx, column=7, value=tx.get('operations', '')).border = thin_border
        ws.cell(row=row_idx, column=8, value=tx.get('description', '')).border = thin_border

        amount_cell = ws.cell(row=row_idx, column=9, value=tx.get('amount', 0))
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = thin_border
        amount_cell.alignment = Alignment(horizontal="right")

        ws.cell(row=row_idx, column=10, value=tx.get('currency', '')).border = thin_border

        fee_cell = ws.cell(row=row_idx, column=11, value=tx.get('bank_fee', 0))
        fee_cell.number_format = '#,##0.00'
        fee_cell.border = thin_border
        fee_cell.alignment = Alignment(horizontal="right")

        ws.cell(row=row_idx, column=12, value=tx.get('entry_type', '')).border = thin_border
        ws.cell(row=row_idx, column=13, value=tx.get('status_display', '')).border = thin_border
        ws.cell(row=row_idx, column=14, value=tx.get('created_at', '')).border = thin_border
    
    total_row = len(transactions) + 2
    ws.cell(row=total_row, column=8, value="TOTAL:")
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=total_amount)
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    ws.cell(row=total_row, column=9).alignment = Alignment(horizontal="right")

    col_widths = [12, 18, 18, 20, 18, 22, 12, 35, 12, 10, 12, 10, 12, 16]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=transacciones_{timestamp}.xlsx'
    return response


def export_to_pdf(transactions, filters):
    if not REPORTLAB_AVAILABLE:
        return JsonResponse({'status': 'error', 'message': 'Librería reportlab no disponible'}, status=500)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        bold=True,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=10,
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=20,
    )
    
    filter_style = ParagraphStyle(
        'FilterInfo',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666'),
        spaceAfter=5,
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=8,
        bold=True,
        textColor=colors.white,
    )
    
    elements = []
    
    elements.append(Paragraph("Reporte de Transacciones Bancarias", title_style))

    elements.append(Paragraph(f"Fecha de generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    
    filter_lines = []
    if filters.get('date_range'):
        filter_lines.append(f"Rango de fechas: {filters['date_range']}")
    if filters.get('date_from'):
        filter_lines.append(f"Desde: {filters['date_from']}")
    if filters.get('date_to'):
        filter_lines.append(f"Hasta: {filters['date_to']}")
    if filters.get('bank'):
        filter_lines.append(f"Banco: {filters['bank']}")
    if filters.get('status'):
        filter_lines.append(f"Estado: {filters['status']}")
    if filters.get('entry_type'):
        filter_lines.append(f"Tipo: {filters['entry_type']}")
    if filters.get('operation_type'):
        filter_lines.append(f"Tipo de operación: {filters['operation_type']}")
    if filters.get('currency'):
        filter_lines.append(f"Moneda: {filters['currency']}")
    if filters.get('amount_min') or filters.get('amount_max'):
        min_amt = filters.get('amount_min', '0')
        max_amt = filters.get('amount_max', 'Sin límite')
        filter_lines.append(f"Monto: {min_amt} - {max_amt}")
    
    if filter_lines:
        elements.append(Paragraph("Filtros aplicados:", ParagraphStyle('Bold', parent=styles['Normal'], bold=True, fontSize=9, spaceAfter=3)))
        for fline in filter_lines:
            elements.append(Paragraph(fline, filter_style))
    
    elements.append(Spacer(1, 10))
    
    headers = ['Fecha', 'Ref. Corriente', 'Ref. Original', 'Banco', 'Monto', 'Moneda', 'Tipo', 'Estado', 'Oficina']
    col_widths = [50, 70, 50, 80, 70, 40, 40, 50, 65]
    
    table_data = [headers]
    
    total_amount = 0
    max_rows = 30
    current_page = 1
    
    for idx, tx in enumerate(transactions):
        total_amount += tx.get('amount', 0)
        
        row = [
            tx.get('date', ''),
            tx.get('reference', ''),
            tx.get('original_reference', ''),
            tx.get('bank', ''),
            f"{tx.get('amount', 0):,.2f}",
            tx.get('currency', ''),
            tx.get('entry_type', ''),
            tx.get('status_display', ''),
            tx.get('office', ''),
        ]
        table_data.append(row)
        
        if len(table_data) - 1 >= max_rows and idx < len(transactions) - 1:
            t = Table(table_data, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            elements.append(t)
            elements.append(Paragraph(f"Página {current_page}", ParagraphStyle('PageNum', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)))
            elements.append(PageBreak())
            
            table_data = [headers]
            current_page += 1
    
    if table_data:
        total_row = [['TOTAL', '', '', '', f"{total_amount:,.2f}", '', '', '', '']]
        
        t = Table(table_data + total_row, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 8),
        ]))
        elements.append(t)
    
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total de registros: {len(transactions)}", ParagraphStyle('Total', parent=styles['Normal'], fontSize=9, bold=True)))
    elements.append(Paragraph(f"Monto total: {total_amount:,.2f}", ParagraphStyle('Total', parent=styles['Normal'], fontSize=9, bold=True)))
    
    doc.build(elements)

    buffer.seek(0)
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=transacciones_{timestamp}.pdf'
    return response